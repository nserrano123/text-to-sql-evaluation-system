#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador del dataset de evaluacion MULTITURNO para el sistema Text-to-SQL.

Objetivo del dataset
--------------------
Evaluar la configuracion foco `RAG+Memoria+Intension+DIN+DEA` en escenarios de
conversacion, midiendo dos capacidades que la exactitud de un solo turno NO
captura:

  1. GATING de intencion: detectar *cuando* debe ejecutarse SQL y cuando NO
     (saludos, cortesias, preguntas meta / fuera de dominio -> NO_SQL;
      peticiones ambiguas sin entidad -> CLARIFY).
  2. MEMORIA de contexto: cuando SI corresponde ejecutar SQL en un turno de
     seguimiento, construir la consulta a partir de la entidad y el resultado
     mencionados en turnos anteriores (p. ej. "dime a que tipo pertenece").

El orden de los turnos es significativo: un turno de seguimiento solo tiene
sentido despues del turno del que depende (columna `depends_on_turn`).

Fuente de las preguntas semilla
-------------------------------
Las preguntas SQL semilla se toman de las filas realmente evaluadas bajo la
configuracion foco en `dataset_sql_40_med_complex.xlsx`
(hoja "Medianas_y_Complejas"), de modo que se anclan al esquema real del ERP.

Salida
------
`analisis_pruebas/preguntas_multiturno.csv`  (UTF-8 con BOM, apto para Excel)
Una fila por turno. 100 conversaciones, de 5 a 10 turnos cada una.
"""

import os
import re
import csv
import random

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "dataset_sql_40_med_complex.xlsx")
SHEET = "Medianas_y_Complejas"
OUTDIR = os.path.join(BASE, "analisis_pruebas")
OUTCSV = os.path.join(OUTDIR, "preguntas_multiturno.csv")

FOCUS = "[Rag],[Memoria],[Intension],[Din],[Dea]"

N_CONVERSACIONES = 100
TURNOS_MIN = 5
TURNOS_MAX = 10
SEED = 42  # reproducibilidad

# --------------------------------------------------------------------------- #
# Vocabulario de dominio: tabla -> (sujeto, plural, atributo de "tipo")
# --------------------------------------------------------------------------- #
TABLE_SUBJECT = {
    "inventory.product":            ("producto", "productos", "tipo de producto"),
    "inventory.product_type":       ("tipo de producto", "tipos de producto", "clasificacion"),
    "inventory.warehouse":          ("bodega", "bodegas", "sucursal"),
    "core.branch_office":           ("sucursal", "sucursales", "zona"),
    "core.category":                ("categoria", "categorias", "grupo"),
    "core.cost_center":             ("centro de costo", "centros de costo", "area"),
    "core.operation":               ("operacion", "operaciones", "tipo de operacion"),
    "accounting.account":           ("cuenta contable", "cuentas contables", "clase financiera"),
    "accounting.movement":          ("movimiento contable", "movimientos contables", "tipo de movimiento"),
    "accounting.movement_tax_detail": ("detalle de impuesto", "detalles de impuesto", "tipo de impuesto"),
    "treasury.bank_account":        ("cuenta bancaria", "cuentas bancarias", "banco"),
    "treasury.payment_method":      ("metodo de pago", "metodos de pago", "tipo de pago"),
    "treasury.payment_schedule":    ("programacion de pago", "programaciones de pago", "estado"),
    "account_payable.advance":      ("anticipo a proveedor", "anticipos a proveedor", "estado"),
    "account_payable.commercial_condition": ("condicion comercial", "condiciones comerciales", "tipo"),
    "account_receivable.customer_detail": ("cliente", "clientes", "segmento"),
    "account_receivable.discount_type": ("tipo de descuento", "tipos de descuento", "categoria"),
    "account_receivable.commercial_condition": ("condicion comercial", "condiciones comerciales", "tipo"),
    "national_transport.route":     ("ruta", "rutas", "zona"),
    "national_transport.remittance": ("remesa", "remesas", "estado"),
    "national_transport.billing_return": ("devolucion de facturacion", "devoluciones de facturacion", "estado"),
    "order.order":                  ("orden", "ordenes", "estado"),
    "tax.tax":                      ("impuesto", "impuestos", "tipo de impuesto"),
    "pos.closed_order_detail":      ("venta", "ventas", "punto de venta"),
    "health.treatment_item":        ("item de tratamiento", "items de tratamiento", "tipo de tratamiento"),
}


def subject_of(table):
    if table in TABLE_SUBJECT:
        return TABLE_SUBJECT[table]
    base = table.split(".")[-1].replace("_", " ")
    return (base, base + "s", "tipo")


# --------------------------------------------------------------------------- #
# Plantillas de turnos de seguimiento
#   text: usa {subject}, {subject_pl}, {attr}
#   action: EXECUTE_SQL | NO_SQL | CLARIFY
#   needs_ctx: si depende de un turno anterior
# --------------------------------------------------------------------------- #
CTX_SQL = [
    ("ctx_tipo",     "Y dime a que {attr} pertenece.",
     "Resolver en memoria la entidad ({subject}) del turno {dep} y devolver su {attr}."),
    ("ctx_top5",     "De ese resultado, muestrame solo los primeros 5.",
     "Reutilizar el resultado del turno {dep} aplicando LIMIT 5."),
    ("ctx_orden",    "Ordenalo de mayor a menor.",
     "Aplicar ORDER BY descendente sobre el resultado del turno {dep}."),
    ("ctx_max",      "Cual de esos tiene el valor mas alto?",
     "Sobre el resultado del turno {dep}, devolver el registro con el maximo."),
    ("ctx_total",    "Cuantos {subject_pl} dio ese resultado en total?",
     "COUNT sobre el conjunto del turno {dep}."),
    ("ctx_activos",  "Ahora dejame solo los que esten activos.",
     "Anadir condicion is_active / audit_status <> 'D' sobre el turno {dep}."),
    ("ctx_mes",      "Dame lo mismo pero solo de lo creado el ultimo mes.",
     "Anadir filtro temporal sobre created_at respecto al turno {dep}."),
    ("ctx_promedio", "Y en vez de la suma, dame el promedio.",
     "Cambiar SUM por AVG respecto al turno {dep}."),
    ("ctx_codigo",   "Agregale tambien el codigo a ese listado.",
     "Anadir la columna code al SELECT del turno {dep}."),
    ("ctx_detalle",  "Muestrame el detalle del primer {subject} de esa lista.",
     "Seleccionar el detalle del primer registro del turno {dep}."),
    ("ctx_nombre",   "De esos, cuales son sus nombres?",
     "Proyectar la columna name de las entidades del turno {dep}."),
]

NO_SQL = [
    ("ns_gracias",   "Listo, muchas gracias, justo lo que necesitaba.",
     "Cortesia de cierre: responder cordialmente SIN ejecutar SQL."),
    ("ns_ok",        "Perfecto, entendido.",
     "Confirmacion breve: no requiere consulta a la base de datos."),
    ("ns_saludo",    "Hola, buenas, como vas?",
     "Saludo conversacional: no debe disparar SQL."),
    ("ns_capacidad", "Que otras cosas puedes consultar de mi empresa?",
     "Pregunta meta sobre capacidades: responder sin ejecutar SQL."),
    ("ns_identidad", "Y tu quien eres exactamente?",
     "Pregunta de identidad del asistente: no requiere SQL."),
    ("ns_offtopic",  "Oye y me recomiendas algo para almorzar hoy?",
     "Fuera de dominio: no debe ejecutar SQL."),
    ("ns_explica",   "Me puedes explicar que significa ese dato?",
     "Explicacion conceptual desde la documentacion del esquema (RAG), SIN ejecutar SQL."),
    ("ns_opinion",   "Ese numero te parece normal para un negocio como el mio?",
     "Solicitud de opinion: responder sin ejecutar SQL."),
]

CLARIFY = [
    ("cl_vago",      "Dame los datos.",
     "Peticion sin entidad ni metrica: el sistema debe PEDIR ACLARACION, no ejecutar SQL."),
    ("cl_otro",      "Y lo otro tambien, porfa.",
     "Referencia ambigua sin antecedente resoluble: pedir aclaracion."),
    ("cl_siempre",   "Muestrame lo de siempre.",
     "Peticion ambigua sin contexto suficiente: pedir aclaracion."),
    ("cl_comparar",  "Comparalo.",
     "Comparacion sin segundo termino explicito: pedir aclaracion."),
    ("cl_eso",       "Y eso como va?",
     "Pregunta vaga sin referente claro: pedir aclaracion."),
]


# --------------------------------------------------------------------------- #
# Carga de semillas
# --------------------------------------------------------------------------- #
def clean_sql(s):
    s = re.sub(r"\s+", " ", str(s)).strip()
    return s


def parse_table(sql):
    m = re.search(r'FROM\s+"?(\w+)"?\."?(\w+)"?', str(sql), re.I)
    if m:
        return f"{m.group(1)}.{m.group(2)}"
    return None


def load_seeds():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    hdr = [c.value for c in ws[1]]
    seen = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, r))
        if (d.get("chatInput") and d.get("Parametros") == FOCUS
                and str(d.get("EVALUADO")) == "1"):
            q = d["chatInput"].strip()
            if q in seen:
                continue
            table = parse_table(d.get("SQL"))
            if not table:
                continue
            seen[q] = {
                "question": q,
                "sql": clean_sql(d.get("SQL")),
                "table": table,
                "clasif": d.get("Clasificacion"),
            }
    seeds = list(seen.values())
    # asignar id estable
    for i, s in enumerate(seeds, 1):
        s["seed_id"] = f"S{i:04d}"
    return seeds


# --------------------------------------------------------------------------- #
# Construccion de conversaciones
# --------------------------------------------------------------------------- #
def realize(template, subject_tuple, dep):
    subject, subject_pl, attr = subject_tuple
    code, text, note = template
    fmt = dict(subject=subject, subject_pl=subject_pl, attr=attr, dep=dep)
    return code, text.format(**fmt), note.format(**fmt)


def build_plan(n_followups, rnd):
    """Lista de roles de seguimiento garantizando el mix pedagogico."""
    plan = ["ctx", "ctx", "nosql"]          # minimo: 2 contextuales + 1 no-sql
    if rnd.random() < 0.55:
        plan.append("clarify")
    if rnd.random() < 0.35:
        plan.append("switch")               # cambio de tema (nueva semilla SQL)
    while len(plan) < n_followups:
        plan.append(rnd.choice(["ctx", "ctx", "nosql", "clarify"]))
    plan = plan[:n_followups]
    rnd.shuffle(plan)
    # 'switch' no debe ser el primer ni el ultimo seguimiento
    if plan and plan[0] == "switch":
        plan[0], plan[-1] = plan[-1], plan[0]
    if len(plan) > 1 and plan[-1] == "switch":
        # mover el switch a una posicion intermedia
        mid = len(plan) // 2
        plan[mid], plan[-1] = plan[-1], plan[mid]
    return plan


def build_conversation(cid, seeds, rnd):
    sid = f"MT{cid:03d}"
    total = rnd.randint(TURNOS_MIN, TURNOS_MAX)
    greeting = rnd.random() < 0.30

    rows = []
    tindex = 0

    def add(action, needs_ctx, dep, subject_entity, msg, gold_sql, note, ttype, seed_id):
        nonlocal tindex
        tindex += 1
        rows.append({
            "conversation_id": cid,
            "session_id": sid,
            "turn_index": tindex,
            "expected_action": action,
            "requires_context": "Si" if needs_ctx else "No",
            "depends_on_turn": dep if dep else "",
            "subject_entity": subject_entity,
            "user_message": msg,
            "gold_sql": gold_sql,
            "expected_behavior": note,
            "turn_type": ttype,
            "seed_id": seed_id,
        })
        return tindex

    # turno de saludo opcional (NO_SQL)
    if greeting:
        add("NO_SQL", False, "", "", "Hola, buenos dias.",
            "", "Saludo inicial: no debe disparar SQL.", "ns_saludo", "")

    # primera semilla SQL (ancla)
    seed = rnd.choice(seeds)
    subj = subject_of(seed["table"])
    active_subject = subj
    active_sql_turn = add("EXECUTE_SQL", False, "", subj[0],
                          seed["question"], seed["sql"],
                          f"Consulta inicial sobre {subj[0]} (tabla {seed['table']}). "
                          "Debe ejecutar SQL.", "seed_sql", seed["seed_id"])

    n_followups = total - tindex
    if n_followups < 3:
        n_followups = 3
    plan = build_plan(n_followups, rnd)

    used_ctx, used_ns, used_cl = set(), set(), set()
    # si ya hubo saludo inicial, no repetir un saludo a mitad de conversacion
    nosql_pool = [t for t in NO_SQL if not (greeting and t[0] == "ns_saludo")]

    def pick(pool, used):
        opts = [t for t in pool if t[0] not in used]
        if not opts:
            used.clear()
            opts = list(pool)
        t = rnd.choice(opts)
        used.add(t[0])
        return t

    for role in plan:
        if role == "switch":
            s2 = rnd.choice(seeds)
            subj2 = subject_of(s2["table"])
            active_subject = subj2
            active_sql_turn = add(
                "EXECUTE_SQL", False, "", subj2[0],
                "Ahora, cambiando de tema: " + s2["question"][0].lower() + s2["question"][1:],
                s2["sql"],
                f"Cambio de tema a {subj2[0]} (tabla {s2['table']}). Debe ejecutar SQL "
                "y reemplazar la entidad activa en memoria.", "switch_sql", s2["seed_id"])
        elif role == "ctx":
            code, msg, note = realize(pick(CTX_SQL, used_ctx), active_subject, active_sql_turn)
            add("EXECUTE_SQL", True, active_sql_turn, active_subject[0],
                msg, "", note, code, "")
        elif role == "nosql":
            t = pick(nosql_pool, used_ns)
            code, msg, note = realize(t, active_subject, active_sql_turn)
            # ns_explica depende del contexto (se refiere a un dato mostrado)
            needs = code in ("ns_explica",)
            add("NO_SQL", needs, active_sql_turn if needs else "",
                active_subject[0] if needs else "", msg, "", note, code, "")
        elif role == "clarify":
            code, msg, note = realize(pick(CLARIFY, used_cl), active_subject, active_sql_turn)
            add("CLARIFY", False, "", "", msg, "", note, code, "")

    return rows


# --------------------------------------------------------------------------- #
def main():
    rnd = random.Random(SEED)
    seeds = load_seeds()
    print(f"Preguntas semilla (config foco): {len(seeds)}")

    all_rows = []
    for cid in range(1, N_CONVERSACIONES + 1):
        all_rows.extend(build_conversation(cid, seeds, rnd))

    os.makedirs(OUTDIR, exist_ok=True)
    cols = ["conversation_id", "session_id", "turn_index", "expected_action",
            "requires_context", "depends_on_turn", "subject_entity",
            "user_message", "gold_sql", "expected_behavior", "turn_type", "seed_id"]
    with open(OUTCSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(all_rows)

    # resumen
    from collections import Counter
    acc = Counter(r["expected_action"] for r in all_rows)
    turns = Counter(r["conversation_id"] for r in all_rows)
    print(f"Conversaciones: {N_CONVERSACIONES}")
    print(f"Turnos totales: {len(all_rows)}")
    print(f"Turnos/conv: min={min(turns.values())} max={max(turns.values())} "
          f"prom={len(all_rows)/N_CONVERSACIONES:.1f}")
    print("Distribucion de expected_action:")
    for k, v in acc.most_common():
        print(f"  {k:12s} {v:4d}  ({v*100/len(all_rows):.1f}%)")
    print("CSV ->", os.path.relpath(OUTCSV, BASE))


if __name__ == "__main__":
    main()
