#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analisis de ablacion Text-to-SQL (ERP Fail Fast).

Lee el dataset crudo de evaluaciones (una fila = una consulta gold ejecutada
bajo una configuracion de componentes) y produce:

  1. Una tabla agregada por configuracion (metricas promedio).
  2. Un conjunto de figuras PNG para la tesis.
  3. Un informe Markdown (analisis_pruebas.md) que embebe las figuras.

Fuente de datos:
  archivo : dataset_sql_40_med_complex.xlsx
  hoja    : "Medianas_y_Complejas"
  filtro  : EVALUADO == 1   (filas efectivamente calificadas)

La columna "Parametros" identifica la configuracion, p. ej.:
  [Rag],[Memoria],[Intension],[Din],[Dea]

Metricas por fila (0..1 salvo VAL_TTL en segundos):
  VAL_SELECT / VAL_WHERE / VAL_GROUP / VAL_ORDER / VAL_KEYWORDS  -> F1 Component Matching
  VAL_EXECUTION      -> Execution Accuracy (1 = resultado correcto)
  VAL_ERRORES        -> SIN ERRORES (robustez CON autocorreccion)
  SIN_AUTO_CORRECCION-> correcto al primer intento (SIN autocorreccion)
  VAL_TTL            -> Time-To-Answer en segundos
"""

import os
import statistics as st
from collections import defaultdict, OrderedDict

import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import PercentFormatter

# --------------------------------------------------------------------------- #
# Configuracion general
# --------------------------------------------------------------------------- #
BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "dataset_sql_40_med_complex.xlsx")
SHEET = "Medianas_y_Complejas"
OUTDIR = os.path.join(BASE, "analisis_pruebas")
FIGDIR = os.path.join(OUTDIR, "figs")
os.makedirs(FIGDIR, exist_ok=True)

# Paleta categorica validada (dataviz skill, modo claro / impresion)
C_BLUE   = "#2a78d6"
C_GREEN  = "#008300"
C_MAGENTA= "#e87ba4"
C_YELLOW = "#eda100"
C_AQUA   = "#1baf7a"
C_ORANGE = "#eb6834"
C_VIOLET = "#4a3aa7"
C_RED    = "#e34948"

INK      = "#0b0b0b"   # texto primario
INK2     = "#52514e"   # texto secundario
MUTED    = "#898781"   # ejes / etiquetas
GRID     = "#e1e0d9"   # rejilla
SURFACE  = "#fcfcfb"   # fondo

# Estilo global coherente para todas las figuras
plt.rcParams.update({
    "figure.facecolor": SURFACE,
    "axes.facecolor": SURFACE,
    "savefig.facecolor": SURFACE,
    "font.family": "DejaVu Sans",
    "font.size": 11,
    "text.color": INK,
    "axes.edgecolor": MUTED,
    "axes.labelcolor": INK2,
    "xtick.color": MUTED,
    "ytick.color": MUTED,
    "axes.grid": True,
    "grid.color": GRID,
    "grid.linewidth": 0.8,
    "axes.spines.top": False,
    "axes.spines.right": False,
})

# Componentes (para el analisis de contribucion marginal)
COMPONENTES = ["Rag", "Memoria", "Intension", "Din", "Dea"]
COMP_LABEL  = {"Rag": "RAG", "Memoria": "Memoria", "Intension": "Intencion",
               "Din": "DIN", "Dea": "DEA"}

# Solo se analizan configuraciones que incluyan RAG: generar SQL sin contexto
# del esquema de la base de datos no tiene sentido en este dominio ERP.
ONLY_RAG = True

# Configuracion foco: aunque no maximiza la exactitud, es la unica que integra
# Memoria (contexto conversacional) e Intencion (deteccion de cuando ejecutar
# SQL), requisitos para el uso multiturno real del sistema.
FOCUS = "[Rag],[Memoria],[Intension],[Din],[Dea]"
FOCUS_LABEL = "RAG+Memoria+Intencion+DIN+DEA"


# --------------------------------------------------------------------------- #
# Carga y agregacion
# --------------------------------------------------------------------------- #
def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def short_label(param):
    """'[Rag],[Din],[Dea]' -> 'RAG+DIN+DEA'."""
    parts = [p.strip("[]") for p in param.split(",")]
    return "+".join(COMP_LABEL.get(p, p) for p in parts)


def has_comp(param, comp):
    return ("[" + comp + "]") in param


def load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    hdr = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, r))
        if d.get("chatInput") and str(d.get("EVALUADO")) == "1":
            if ONLY_RAG and not has_comp(d.get("Parametros", ""), "Rag"):
                continue
            rows.append(d)
    # COMPARABILIDAD: todas las configuraciones se agregan sobre el MISMO
    # subconjunto de 53 preguntas gold (la config completa tiene ademas ~1000
    # preguntas extra de un solo turno que inflarian/deformarian sus promedios
    # frente a las demas). Referencia: bloque [Rag],[Din],[Dea] (n=53).
    gold53 = {str(r["chatInput"]).strip() for r in rows
              if r["Parametros"] == "[Rag],[Din],[Dea]"}
    seen = set()
    comp = []
    for r in rows:
        q = str(r["chatInput"]).strip()
        if q not in gold53:
            continue
        key = (r["Parametros"], q)
        if key in seen:
            continue
        seen.add(key)
        comp.append(r)
    return comp


def aggregate(rows):
    """Devuelve OrderedDict config -> dict de metricas promedio, ordenado por EX desc."""
    groups = defaultdict(list)
    for r in rows:
        groups[r["Parametros"]].append(r)

    metric_cols = ["VAL_SELECT", "VAL_WHERE", "VAL_GROUP", "VAL_ORDER",
                   "VAL_KEYWORDS", "VAL_EXECUTION", "VAL_ERRORES",
                   "SIN_AUTO_CORRECCION", "VAL_TTL"]

    agg = {}
    for param, rs in groups.items():
        m = {"config": param, "label": short_label(param), "n": len(rs)}
        for col in metric_cols:
            xs = [num(r.get(col)) for r in rs]
            xs = [x for x in xs if x is not None]
            m[col] = sum(xs) / len(xs) if xs else None
        # Component Matching = promedio de SELECT, WHERE, GROUP, KEYWORDS (como en la tesis)
        cm_parts = [m[c] for c in ("VAL_SELECT", "VAL_WHERE", "VAL_GROUP", "VAL_KEYWORDS")
                    if m[c] is not None]
        m["CM"] = sum(cm_parts) / len(cm_parts) if cm_parts else None
        agg[param] = m

    ordered = OrderedDict(
        sorted(agg.items(), key=lambda kv: -(kv[1]["VAL_EXECUTION"] or 0))
    )
    return ordered


# --------------------------------------------------------------------------- #
# Utilidades de dibujo
# --------------------------------------------------------------------------- #
def save(fig, name):
    path = os.path.join(FIGDIR, name)
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    print("  figura ->", os.path.relpath(path, BASE))


# --------------------------------------------------------------------------- #
# Figuras
# --------------------------------------------------------------------------- #
def fig_execution_ranking(agg):
    """Ranking horizontal de Execution Accuracy por configuracion."""
    data = [(m["label"], m["VAL_EXECUTION"]) for m in agg.values()
            if m["VAL_EXECUTION"] is not None]
    data.sort(key=lambda t: t[1])
    labels = [d[0] for d in data]
    vals = [d[1] * 100 for d in data]

    fig, ax = plt.subplots(figsize=(9, 7.2))
    vmax = max(vals)
    colors = []
    for lab, v in zip(labels, vals):
        if lab == FOCUS_LABEL:
            colors.append(C_ORANGE)      # configuracion foco (multiturno)
        elif v == vmax:
            colors.append(C_GREEN)       # mejor EX
        else:
            colors.append(C_BLUE)
    bars = ax.barh(labels, vals, color=colors, height=0.72, zorder=3)
    for b, v in zip(bars, vals):
        ax.text(v + 0.8, b.get_y() + b.get_height() / 2, f"{v:.1f}%",
                va="center", ha="left", fontsize=9, color=INK2)
    # resaltar la etiqueta de la config foco
    for tick in ax.get_yticklabels():
        if tick.get_text() == FOCUS_LABEL:
            tick.set_color(C_ORANGE)
            tick.set_fontweight("bold")
    ax.set_xlim(0, 100)
    ax.xaxis.set_major_formatter(PercentFormatter())
    ax.set_xlabel("Execution Accuracy")
    ax.set_title("Exactitud de ejecucion (solo configuraciones con RAG)",
                 color=INK, fontsize=13, pad=12, loc="left")
    # leyenda de resaltados
    from matplotlib.patches import Patch
    ax.legend(handles=[Patch(color=C_GREEN, label="Mejor EX"),
                       Patch(color=C_ORANGE, label="Config. foco (multiturno)")],
              frameon=False, loc="lower right", fontsize=9)
    ax.grid(axis="y", visible=False)
    ax.tick_params(axis="y", labelsize=8.5)
    save(fig, "fig1_execution_ranking.png")


def fig_cm_vs_ex(agg):
    """Dispersion: Component Matching vs Execution Accuracy (disociacion)."""
    xs, ys, labs = [], [], []
    for m in agg.values():
        if m["CM"] is not None and m["VAL_EXECUTION"] is not None:
            xs.append(m["CM"] * 100)
            ys.append(m["VAL_EXECUTION"] * 100)
            labs.append(m["label"])

    fig, ax = plt.subplots(figsize=(8.5, 6.5))
    ax.scatter(xs, ys, s=70, color=C_BLUE, edgecolor="white", linewidth=1.2, zorder=3)

    # Etiquetar solo casos narrativos: mejor, config foco, CM alto / EX bajo y RAG solo
    interesantes = {"RAG+DIN+DEA", "RAG+Intencion+DIN+DEA",
                    FOCUS_LABEL, "RAG"}
    for x, y, l in zip(xs, ys, labs):
        if l in interesantes:
            ax.annotate(l, (x, y), textcoords="offset points", xytext=(6, 6),
                        fontsize=8.5, color=INK)
    ax.set_xlabel("Component Matching (F1 promedio)")
    ax.set_ylabel("Execution Accuracy")
    ax.xaxis.set_major_formatter(PercentFormatter())
    ax.yaxis.set_major_formatter(PercentFormatter())
    ax.set_title("Calidad estructural vs. correccion funcional",
                 color=INK, fontsize=13, pad=12, loc="left")
    save(fig, "fig2_cm_vs_ex.png")


def fig_cm_breakdown(agg):
    """Component Matching desagregado (SELECT/WHERE/GROUP/KEYWORDS) para configs clave."""
    orden = ["RAG+DIN+DEA", "RAG+Memoria+Intencion+DIN+DEA", "RAG+Dea",
             "RAG+Intencion+DIN+DEA", "RAG+Intencion+DEA"]
    by_label = {m["label"]: m for m in agg.values()}
    sel = [by_label[l] for l in orden if l in by_label]

    comps = [("VAL_SELECT", "SELECT", C_BLUE),
             ("VAL_WHERE", "WHERE", C_GREEN),
             ("VAL_GROUP", "GROUP BY", C_MAGENTA),
             ("VAL_KEYWORDS", "KEYWORDS", C_YELLOW)]

    import numpy as np
    x = np.arange(len(sel))
    w = 0.20
    fig, ax = plt.subplots(figsize=(9.5, 5.6))
    for i, (col, name, color) in enumerate(comps):
        vals = [(m[col] or 0) * 100 for m in sel]
        ax.bar(x + (i - 1.5) * w, vals, w, label=name, color=color, zorder=3)
    ax.set_xticks(x)
    ax.set_xticklabels([m["label"] for m in sel], fontsize=8.5, rotation=12, ha="right")
    ax.set_ylim(0, 105)
    ax.yaxis.set_major_formatter(PercentFormatter())
    ax.set_ylabel("F1-score")
    ax.grid(axis="x", visible=False)
    ax.set_title("Component Matching desagregado por clausula SQL",
                 color=INK, fontsize=13, pad=12, loc="left")
    ax.legend(frameon=False, ncol=4, loc="lower center",
              bbox_to_anchor=(0.5, -0.28), fontsize=9)
    save(fig, "fig3_cm_breakdown.png")


def fig_autocorreccion_gap(agg):
    """SIN ERRORES (con autocorreccion) vs primer intento, ordenado por brecha."""
    import numpy as np
    data = []
    for m in agg.values():
        if m["VAL_ERRORES"] is not None and m["SIN_AUTO_CORRECCION"] is not None:
            data.append((m["label"], m["VAL_ERRORES"] * 100,
                         m["SIN_AUTO_CORRECCION"] * 100))
    data.sort(key=lambda t: (t[1] - t[2]))  # menor brecha arriba
    labels = [d[0] for d in data]
    con_ac = [d[1] for d in data]
    primer = [d[2] for d in data]

    y = np.arange(len(labels))
    fig, ax = plt.subplots(figsize=(9.5, 8.5))
    ax.hlines(y, primer, con_ac, color=GRID, linewidth=2.2, zorder=1)
    ax.scatter(con_ac, y, s=52, color=C_BLUE, label="SIN ERRORES (con autocorreccion)",
               zorder=3, edgecolor="white", linewidth=0.8)
    ax.scatter(primer, y, s=52, color=C_ORANGE, label="Correcto al primer intento",
               zorder=3, edgecolor="white", linewidth=0.8)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=8)
    ax.set_xlim(0, 105)
    ax.xaxis.set_major_formatter(PercentFormatter())
    ax.grid(axis="y", visible=False)
    ax.set_xlabel("Proporcion de consultas exitosas")
    ax.set_title("Dependencia de la autocorreccion por configuracion",
                 color=INK, fontsize=13, pad=12, loc="left")
    ax.legend(frameon=False, ncol=2, loc="lower center",
              bbox_to_anchor=(0.5, -0.14), fontsize=9)
    save(fig, "fig4_autocorreccion_gap.png")


def fig_contribucion_marginal(rows):
    """Contribucion marginal media de cada componente sobre Execution Accuracy.

    Compara la EX media de las configuraciones que incluyen el componente
    frente a las que no lo incluyen (promedio a nivel de configuracion).
    """
    import numpy as np
    groups = defaultdict(list)
    for r in rows:
        groups[r["Parametros"]].append(r)
    # EX por configuracion
    ex_by_cfg = {}
    for param, rs in groups.items():
        xs = [num(r.get("VAL_EXECUTION")) for r in rs]
        xs = [x for x in xs if x is not None]
        if xs:
            ex_by_cfg[param] = sum(xs) / len(xs)

    deltas = []
    for comp in COMPONENTES:
        con = [v for p, v in ex_by_cfg.items() if has_comp(p, comp)]
        sin = [v for p, v in ex_by_cfg.items() if not has_comp(p, comp)]
        if con and sin:
            deltas.append((COMP_LABEL[comp],
                           (st.mean(con) - st.mean(sin)) * 100,
                           len(con), len(sin)))
    deltas.sort(key=lambda t: t[1])

    labels = [d[0] for d in deltas]
    vals = [d[1] for d in deltas]
    colors = [C_BLUE if v >= 0 else C_RED for v in vals]

    fig, ax = plt.subplots(figsize=(8, 4.8))
    bars = ax.barh(labels, vals, color=colors, height=0.6, zorder=3)
    for b, v in zip(bars, vals):
        ax.text(v + (0.4 if v >= 0 else -0.4), b.get_y() + b.get_height() / 2,
                f"{v:+.1f} pp", va="center",
                ha="left" if v >= 0 else "right", fontsize=9, color=INK2)
    ax.axvline(0, color=MUTED, linewidth=1)
    ax.set_xlabel("Delta medio en Execution Accuracy (puntos porcentuales)")
    ax.grid(axis="y", visible=False)
    ax.set_title("Contribucion marginal media de cada componente",
                 color=INK, fontsize=13, pad=12, loc="left")
    lim = max(abs(min(vals)), abs(max(vals))) * 1.35
    ax.set_xlim(-lim, lim)
    save(fig, "fig5_contribucion_marginal.png")


def fig_tta_tradeoff(agg):
    """Compromiso: Time-To-Answer vs Execution Accuracy."""
    xs, ys, labs, ns = [], [], [], []
    for m in agg.values():
        if m["VAL_TTL"] and m["VAL_EXECUTION"] is not None:
            xs.append(m["VAL_TTL"])
            ys.append(m["VAL_EXECUTION"] * 100)
            labs.append(m["label"])
    if not xs:
        print("  (sin datos de VAL_TTL, se omite fig6)")
        return
    fig, ax = plt.subplots(figsize=(8.5, 6.2))
    ax.scatter(xs, ys, s=70, color=C_VIOLET, edgecolor="white", linewidth=1.1, zorder=3)
    interesantes = {"RAG+DIN+DEA", "RAG+Memoria+Intencion+DIN+DEA", "Intencion+DIN+DEA"}
    for x, y, l in zip(xs, ys, labs):
        if l in interesantes:
            ax.annotate(l, (x, y), textcoords="offset points", xytext=(6, 5),
                        fontsize=8.5, color=INK)
    ax.set_xlabel("Time-To-Answer medio (segundos)")
    ax.set_ylabel("Execution Accuracy")
    ax.yaxis.set_major_formatter(PercentFormatter())
    ax.set_title("Compromiso entre latencia y exactitud",
                 color=INK, fontsize=13, pad=12, loc="left")
    save(fig, "fig6_tta_tradeoff.png")


# --------------------------------------------------------------------------- #
# Markdown
# --------------------------------------------------------------------------- #
def pct(x):
    return f"{x*100:.1f}\\%".replace("\\", "") if x is not None else "n/d"


def write_markdown(agg, rows):
    lines = []
    A = lines.append
    n_cfg = len(agg)
    total_eval = len(rows)
    best = next(iter(agg.values()))
    worst = list(agg.values())[-1]

    A("# Analisis de pruebas del sistema Text-to-SQL (estudio de ablacion)\n")
    A("> Generado automaticamente por `generar_analisis_ablacion.py` a partir de "
      "`dataset_sql_40_med_complex.xlsx` (hoja **Medianas_y_Complejas**, filas con "
      "`EVALUADO = 1`).\n")
    A("## 1. Alcance de los datos\n")
    A("- **Solo se consideran configuraciones que incluyen RAG.** Generar SQL sin "
      "recuperacion del contexto del esquema no tiene sentido en este dominio ERP: "
      "el modelo no puede resolver tablas ni columnas de forma fiable sin ese anclaje. "
      "Las combinaciones sin RAG se excluyen del analisis.")
    A(f"- Configuraciones con RAG efectivamente evaluadas: **{n_cfg}**.")
    A(f"- Total de ejecuciones calificadas: **{total_eval}**.")
    A("- Cada configuracion se identifica por la combinacion de componentes activos "
      "`[Rag]`, `[Memoria]`, `[Intension]`, `[Din]`, `[Dea]`.")
    A("- Metricas por consulta: Component Matching (F1 de SELECT/WHERE/GROUP BY/KEYWORDS), "
      "Execution Accuracy (EX), SIN ERRORES (robustez con autocorreccion), "
      "correcto al primer intento (sin autocorreccion) y Time-To-Answer (TTA).\n")

    A("## 2. Tabla resumen por configuracion\n")
    A("Ordenada por Execution Accuracy descendente.\n")
    A("| # | Configuracion | n | CM (F1) | EX | SIN ERRORES | 1er intento | TTA (s) |")
    A("|---|---------------|---|---------|----|-------------|-------------|---------|")
    for i, m in enumerate(agg.values(), 1):
        tta = f"{m['VAL_TTL']:.1f}" if m["VAL_TTL"] else "n/d"
        A(f"| {i} | {m['label']} | {m['n']} | {pct(m['CM'])} | {pct(m['VAL_EXECUTION'])} "
          f"| {pct(m['VAL_ERRORES'])} | {pct(m['SIN_AUTO_CORRECCION'])} | {tta} |")
    A("")

    focus = next((m for m in agg.values() if m["label"] == FOCUS_LABEL), None)

    A("## 3. Exactitud de ejecucion\n")
    A(f"![Ranking de Execution Accuracy](figs/fig1_execution_ranking.png)\n")
    A(f"La mejor configuracion en exactitud de un solo turno es **{best['label']}** con "
      f"**{pct(best['VAL_EXECUTION'])}** de Execution Accuracy. Sin embargo, el analisis "
      "posterior se centra en la configuracion completa por razones de aplicabilidad "
      "real (ver seccion siguiente).\n")

    if focus:
        A(f"## 4. Configuracion foco: {FOCUS_LABEL}\n")
        A("Aunque en exactitud de un solo turno esta configuracion "
          f"(**{pct(focus['VAL_EXECUTION'])}** de EX) no es la mejor, es la unica que "
          "integra los dos componentes imprescindibles para el uso real del sistema:")
        A("- **Memoria**: preserva el contexto conversacional, condicion necesaria para "
          "resolver preguntas de seguimiento (p. ej. *\"dime a que tipo pertenece\"* "
          "refiriendose a una entidad mencionada en un turno anterior).")
        A("- **Intencion**: decide *cuando* debe ejecutarse SQL y cuando no, evitando "
          "que mensajes conversacionales, ambiguos o fuera de dominio disparen consultas "
          "innecesarias contra la base de datos.")
        A("")
        A(f"Su perfil de metricas es: CM {pct(focus['CM'])}, EX {pct(focus['VAL_EXECUTION'])}, "
          f"SIN ERRORES {pct(focus['VAL_ERRORES'])}, correcto al primer intento "
          f"{pct(focus['SIN_AUTO_CORRECCION'])}. En exactitud y generacion directa cede "
          "algunos puntos frente a RAG+DIN+DEA: es el costo de anadir los componentes "
          "conversacionales; su justificacion no esta en el turno aislado sino en el "
          "escenario multiturno.")
        A("")
        A("> Por esta razon, la evaluacion multiturno del sistema (dataset "
          "`preguntas_multiturno.csv`) se disena a partir de las preguntas de esta "
          "configuracion: el objetivo es medir si el sistema (i) detecta correctamente "
          "cuando ejecutar SQL y cuando no, y (ii) cuando lo ejecuta, construye la "
          "consulta usando el contexto acumulado de la conversacion.\n")

    A("## 5. Calidad estructural vs. correccion funcional\n")
    A(f"![Component Matching vs Execution Accuracy](figs/fig2_cm_vs_ex.png)\n")
    A("Incluso entre configuraciones con RAG, un Component Matching similar produce "
      "Execution Accuracy muy distintas: la correspondencia estructural es condicion "
      "necesaria pero no suficiente para generar SQL ejecutable. El caso mas ilustrativo "
      "es **RAG+Intencion+DIN+DEA**, con CM elevado (~81,6\\%) pero una EX de solo "
      "64,2\\%: anadir la clasificacion de Intencion sobre RAG+DIN+DEA introduce un "
      "sesgo que degrada la ejecucion pese a mantener la calidad estructural.\n")
    A(f"![Component Matching desagregado](figs/fig3_cm_breakdown.png)\n")
    A("Al desagregar el Component Matching, SELECT, WHERE y KEYWORDS concentran los "
      "valores mas altos, mientras **GROUP BY** es sistematicamente inferior: las "
      "operaciones de agregacion suelen quedar implicitas en el lenguaje natural y "
      "exigen mayor inferencia semantica.\n")

    A("## 6. Robustez y dependencia de la autocorreccion\n")
    A(f"![Brecha de autocorreccion](figs/fig4_autocorreccion_gap.png)\n")
    A("La distancia entre *SIN ERRORES* (con autocorreccion) y *correcto al primer "
      "intento* mide cuanto depende cada configuracion de los ciclos de reparacion. "
      "Configuraciones minimas como **RAG** o **RAG+Intencion** muestran una brecha "
      "amplia: terminan siendo funcionales, pero a costa de multiples iteraciones de "
      "correccion. En cambio, **RAG+DIN+DEA** y la configuracion foco "
      f"**{FOCUS_LABEL}** exhiben una brecha estrecha, es decir, generan SQL correcto "
      "desde el primer intento — propiedad clave para un flujo conversacional fluido.\n")

    A("## 7. Contribucion marginal de cada componente (dentro de RAG)\n")
    A(f"![Contribucion marginal](figs/fig5_contribucion_marginal.png)\n")
    A("Como todas las configuraciones analizadas ya incluyen RAG, este grafico aisla el "
      "aporte marginal de los componentes *adicionales* sobre esa base. **DIN** es el "
      "que mas aporta a la exactitud de ejecucion (~+9,5 pp), seguido de **Memoria** "
      "(~+3,4 pp) y **DEA** (~+0,6 pp). La clasificacion de **Intencion** presenta un "
      "aporte marginal *medio* negativo (~-5,6 pp) sobre la exactitud: su valor no esta "
      "en mejorar la generacion de SQL, sino en **filtrar cuando debe o no ejecutarse "
      "una consulta**, un beneficio que no se refleja en la Execution Accuracy de un "
      "solo turno pero que es esencial en el escenario multiturno.\n")

    if any(m["VAL_TTL"] for m in agg.values()):
        A("## 8. Compromiso latencia / exactitud\n")
        A(f"![Compromiso TTA vs EX](figs/fig6_tta_tradeoff.png)\n")
        A("El TTA permite valorar el costo operativo de cada ganancia de exactitud, "
          "aspecto critico en un entorno ERP donde la latencia condiciona la experiencia "
          "de usuario.\n")

    A("## 9. Sintesis\n")
    A("1. RAG es un requisito de base: sin recuperacion del esquema la generacion de SQL "
      "no es viable, por lo que el analisis se restringe a configuraciones con RAG.")
    A("2. Un alto Component Matching no garantiza SQL ejecutable: hay que medir "
      "explicitamente la ejecucion.")
    A("3. **RAG+DIN+DEA** maximiza la exactitud de un solo turno; **DIN** es el "
      "componente adicional de mayor aporte marginal.")
    A(f"4. La configuracion foco **{FOCUS_LABEL}** se selecciona para el despliegue real "
      "porque anade Memoria e Intencion — los componentes que habilitan el dialogo "
      "multiturno y la deteccion de cuando ejecutar SQL —, aceptando una perdida "
      "moderada de exactitud y de generacion directa frente a RAG+DIN+DEA; la "
      "evaluacion multiturno (759 turnos reales) confirma que esa perdida compra "
      "gating de intencion del 98,4% y resolucion de contexto del 97,4%.")
    A("5. La evaluacion multiturno (dataset `preguntas_multiturno.csv`) mide justamente "
      "esas dos capacidades que la Execution Accuracy de un solo turno no captura.\n")

    path = os.path.join(OUTDIR, "analisis_pruebas.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print("markdown ->", os.path.relpath(path, BASE))


# --------------------------------------------------------------------------- #
def main():
    print("Leyendo:", os.path.relpath(XLSX, BASE))
    rows = load_rows()
    print(f"Filas evaluadas: {len(rows)}")
    agg = aggregate(rows)
    print(f"Configuraciones: {len(agg)}\n")

    print("Generando figuras:")
    fig_execution_ranking(agg)
    fig_cm_vs_ex(agg)
    fig_cm_breakdown(agg)
    fig_autocorreccion_gap(agg)
    fig_contribucion_marginal(rows)
    fig_tta_tradeoff(agg)
    print()
    write_markdown(agg, rows)
    print("\nListo. Revisa la carpeta:", os.path.relpath(OUTDIR, BASE))


if __name__ == "__main__":
    main()
